import xlwt
import openpyxl
import os
import numpy as np

p = "D:\\invoices2.xlsx"
wb2 = openpyxl.load_workbook(p)
sh2 = wb2.active
profit = 0
a = 1
for i in range(2, sh2.max_row ):
    #print(a)
    if (float(sh2.cell(i,9).value) <= float(sh2.cell(i,12).value) ):
        sh2.cell(i,13).value = "z"
    p1 = (float(sh2.cell(i,9).value) - float(sh2.cell(i,12).value))* float(sh2.cell(i,5).value) - int(sh2.cell(i,10).value)
    sh2.cell(i,14).value = p1
    profit += p1
    sh2.cell(i,15).value = profit
    print (p1, "   ", profit)
    #a += 1
print(profit)
wb2.save("invoices2.xlsx")