import xlwt
import openpyxl
import os
import numpy as np

path = "D:\\invoices.xlsx"
all_invoices_obj = openpyxl.load_workbook(path)
sheet_i = all_invoices_obj.active
path = "D:\\aaaa.xlsx"
aaaa_obj = openpyxl.load_workbook(path)
sheet_a = aaaa_obj.active


for i in range(2, sheet_i.max_row + 1):
    shomarande = 1
    for y in range(2, sheet_a.max_row + 1):
        #print(sheet_i.cell(i,8).value ,"    ", sheet_a.cell(y,3).value)
        if (sheet_i.cell(i,2).value == sheet_a.cell(y,1).value):
            #print("a")
            #print(sheet_i.cell(i,8).value ,"    ", sheet_a.cell(y,3).value)
            if (str(sheet_i.cell(i,8).value) == str(sheet_a.cell(y,3).value) ):
                #print("a")
                sheet_i.cell(row = i, column = shomarande + 16 ).value = sheet_a.cell(y, 4).value
                print(sheet_i.cell(row = i, column = shomarande + 16 ).value,  sheet_i.cell(i,2).value, sheet_i.cell(i,8).value)
                shomarande += 1

#sheet_i.cell(row = 1, column = 18 ).value = "chg"
all_invoices_obj.save("invoices2.xlsx")
print("done")