import xlwt
import openpyxl
import os
path = "D:\\invoices.xlsx"
all_invoices_obj = openpyxl.load_workbook(path)
sheet_obj = all_invoices_obj.active

a = sheet_obj.cell(1 , sheet_obj.max_column+1)
#a.value = ""
#print(a)
#print (sheet_obj.max_row)
#print (sheet_obj.max_column)
"""b = []
for i in range(2, sheet_obj.max_row + 1):
    #print(i)
    b.append(sheet_obj.cell(i,11).value)
    #print(sheet_obj.cell(i,11).value)
"""
#print(b)
"""for e in b:
    u = 0
    for t in b:
        if (e == t):
            u +=1
    if (u>1):
        b.remove(e)"""
#print(b)
"""for e in b:
    print(e)
    o = int(input())"""
arr = os.listdir("D:\\kartex")
#print(arr)
#print(len(arr))
#path = os.path.join("D:\\", "review-warehouse1 (1).xlsx")
w = xlwt.Workbook()
worksheet = w.add_sheet("data")
worksheet.write(0,0,"تاریخ")
worksheet.write(0,1,"طرف")
worksheet.write(0,2,"صادره")
worksheet.write(0,3,"فی صادره")
worksheet.write(0,4,"مبلغ صادره")
s = 1
for i in arr:
    path = os.path.join("D:\\", i)
    #print(path)
    wb = openpyxl.load_workbook(path)
    sh_obj = wb.active
    print("folders name: ",i)
    for t in range(2, sh_obj.max_row + 1):
        if(sh_obj.cell(t,4).value == "شرکت فولاد آذر"):
            worksheet.write(s, 0, str(sh_obj.cell(t,1).value))
            worksheet.write(s, 1, str(sh_obj.cell(t,4).value))
            worksheet.write(s, 2, str(sh_obj.cell(t,9).value))
            worksheet.write(s, 3, str(sh_obj.cell(t,10).value))
            worksheet.write(s, 4, str(sh_obj.cell(t,11).value))
            print(sh_obj.cell(t,4).value)
            s += 1
print(s)
    
w.save("aaa.xls")
"""wb = openpyxl.load_workbook("D:\\review-warehouse1 (1).xlsx")
sh_obj = wb.active"""
print("done")































