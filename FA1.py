import xlwt
import openpyxl
import os
import numpy as np

path = "D:\\aaaa.xlsx"
all_invoices_obj = openpyxl.load_workbook(path)
sheet_obj = all_invoices_obj.active

data = np.array([])
d = np.zeros(199)
dates = []
for i in range(2, sheet_obj.max_row + 1):
    data = np.append(data, sheet_obj.cell(i,1).value)
    data = np.append(data, sheet_obj.cell(i,2).value)
    data = np.append(data, sheet_obj.cell(i,3).value)
    data = np.append(data, sheet_obj.cell(i,4).value)
    dates.append(data, sheet_obj.cell(i,1).value)

s = 1
u = 0 
while(s<199):
        for t in dates:
            a = data[u]
            




print(len(data))